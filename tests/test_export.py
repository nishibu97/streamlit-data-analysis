"""データエクスポート機能のテストモジュール"""

import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils.export import (
    export_to_csv,
    export_to_excel,
    export_to_json,
    generate_filename,
    get_mime_type,
)


@pytest.fixture
def sample_dataframe():
    """テスト用のサンプルDataFrame"""
    return pd.DataFrame(
        {
            "回答者ID": [1, 2, 3, 4, 5],
            "年齢層": ["10代", "20代", "30代", "40代", "50代"],
            "サッカー": [5, 4, 3, 2, 1],
            "野球": [3, 4, 5, 4, 3],
            "バスケットボール": [4, 5, 3, 2, 4],
        }
    )


class TestExportToCSV:
    """CSV エクスポート機能のテスト"""

    def test_export_to_csv_returns_bytes(self, sample_dataframe):
        """CSV エクスポートがバイトデータを返すことを確認"""
        result = export_to_csv(sample_dataframe)
        assert isinstance(result, bytes)

    def test_export_to_csv_contains_header(self, sample_dataframe):
        """CSV エクスポート結果にヘッダーが含まれることを確認"""
        result = export_to_csv(sample_dataframe)
        csv_text = result.decode("utf-8-sig")
        assert "回答者ID" in csv_text
        assert "年齢層" in csv_text
        assert "サッカー" in csv_text

    def test_export_to_csv_no_index(self, sample_dataframe):
        """CSV エクスポート結果にインデックスが含まれないことを確認"""
        result = export_to_csv(sample_dataframe)
        csv_text = result.decode("utf-8-sig")
        lines = csv_text.strip().split("\n")
        # データ行数 + ヘッダー行 = 6行
        assert len(lines) == 6


class TestExportToExcel:
    """Excel エクスポート機能のテスト"""

    def test_export_to_excel_returns_bytes(self, sample_dataframe):
        """Excel エクスポートがバイトデータを返すことを確認"""
        result = export_to_excel(sample_dataframe)
        assert isinstance(result, bytes)

    def test_export_to_excel_readable(self, sample_dataframe, tmp_path):
        """エクスポートされた Excel ファイルが読み込み可能であることを確認"""
        result = export_to_excel(sample_dataframe, sheet_name="TestSheet")

        # 一時ファイルに保存して読み込み
        temp_file = tmp_path / "test.xlsx"
        temp_file.write_bytes(result)

        # openpyxl で読み込み確認
        workbook = load_workbook(temp_file)
        assert "TestSheet" in workbook.sheetnames

    def test_export_to_excel_preserves_data(self, sample_dataframe, tmp_path):
        """Excel エクスポートがデータを正しく保持することを確認"""
        result = export_to_excel(sample_dataframe)

        temp_file = tmp_path / "test.xlsx"
        temp_file.write_bytes(result)

        # pandas で読み込んで比較
        df_loaded = pd.read_excel(temp_file)
        pd.testing.assert_frame_equal(df_loaded, sample_dataframe)


class TestExportToJSON:
    """JSON エクスポート機能のテスト"""

    def test_export_to_json_returns_bytes(self, sample_dataframe):
        """JSON エクスポートがバイトデータを返すことを確認"""
        result = export_to_json(sample_dataframe)
        assert isinstance(result, bytes)

    def test_export_to_json_valid_format(self, sample_dataframe):
        """エクスポートされた JSON が有効な形式であることを確認"""
        result = export_to_json(sample_dataframe)
        json_text = result.decode("utf-8")
        parsed = json.loads(json_text)
        assert isinstance(parsed, list)
        assert len(parsed) == 5

    def test_export_to_json_contains_data(self, sample_dataframe):
        """JSON エクスポート結果にデータが含まれることを確認"""
        result = export_to_json(sample_dataframe)
        json_text = result.decode("utf-8")
        parsed = json.loads(json_text)
        first_record = parsed[0]
        assert first_record["回答者ID"] == 1
        assert first_record["年齢層"] == "10代"
        assert first_record["サッカー"] == 5


class TestGenerateFilename:
    """ファイル名生成のテスト"""

    def test_generate_filename_csv(self):
        """CSV ファイル名の生成を確認"""
        filename = generate_filename("test_data", "csv")
        assert filename.startswith("test_data_")
        assert filename.endswith(".csv")

    def test_generate_filename_excel(self):
        """Excel ファイル名の生成を確認"""
        filename = generate_filename("test_data", "excel")
        assert filename.startswith("test_data_")
        assert filename.endswith(".xlsx")

    def test_generate_filename_json(self):
        """JSON ファイル名の生成を確認"""
        filename = generate_filename("test_data", "json")
        assert filename.startswith("test_data_")
        assert filename.endswith(".json")

    def test_generate_filename_contains_timestamp(self):
        """ファイル名にタイムスタンプが含まれることを確認"""
        filename = generate_filename("data", "csv")
        # タイムスタンプ形式（YYYYMMDD_HHMMSS）を含むことを確認
        parts = filename.replace(".csv", "").split("_")
        assert len(parts) >= 3  # data, YYYYMMDD, HHMMSS


class TestGetMimeType:
    """MIME タイプ取得のテスト"""

    def test_get_mime_type_csv(self):
        """CSV の MIME タイプを確認"""
        assert get_mime_type("csv") == "text/csv"

    def test_get_mime_type_excel(self):
        """Excel の MIME タイプを確認"""
        mime_type = get_mime_type("excel")
        assert mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_get_mime_type_json(self):
        """JSON の MIME タイプを確認"""
        assert get_mime_type("json") == "application/json"

    def test_get_mime_type_unknown(self):
        """未知の形式のデフォルト MIME タイプを確認"""
        assert get_mime_type("unknown") == "application/octet-stream"
